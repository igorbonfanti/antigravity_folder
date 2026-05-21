# -*- coding: utf-8 -*-
"""
Kerakoll Scraper — estrae i link PDF delle schede tecniche italiane
dal sito https://it.kerakoll.com/p/<slug> per ogni prodotto del listino Excel.

Output: src/data/products_kerakoll.json
"""

import asyncio
import json
import re
import sys
import unicodedata
from pathlib import Path

import openpyxl
from playwright.async_api import async_playwright, TimeoutError as PlaywrightTimeout

sys.stdout.reconfigure(encoding="utf-8")

# ─── Configurazione ───────────────────────────────────────────────────────────
EXCEL_FILE    = "kerakoll_listino-gen-ita-nov2025.xlsx"
OUTPUT_FILE   = "src/data/products_kerakoll.json"
PROGRESS_FILE = "scrape_progress.json"   # checkpoint per riprendere in caso di interruzione
BASE_URL      = "https://it.kerakoll.com"
CONCURRENCY   = 3    # pagine in parallelo (non esagerare per non essere bloccati)
PAGE_TIMEOUT  = 20000  # ms per caricare la pagina
# ──────────────────────────────────────────────────────────────────────────────


def slugify(text: str) -> str:
    """Converte un nome prodotto nello slug URL usato da Kerakoll."""
    if not text:
        return ""
    # Normalizza unicode → ASCII (accenti → lettere base)
    text = unicodedata.normalize("NFKD", text)
    text = text.encode("ascii", "ignore").decode("ascii")
    # Minuscolo
    text = text.lower()
    # Rimuovi apostrofi e caratteri speciali, sostituisci spazi con trattini
    text = re.sub(r"['\u2019\u2018]", "", text)       # apostrofi
    text = re.sub(r"[^a-z0-9\s\-]", " ", text)        # solo alfanumerici e trattini
    text = re.sub(r"\s+", "-", text.strip())            # spazi → trattini
    text = re.sub(r"-+", "-", text)                     # trattini multipli → uno solo
    return text


def load_excel_products(path: str) -> dict:
    """
    Legge il listino Excel e restituisce un dizionario:
    { cod_prodotto: { nome, linea, categoria, sottocategoria, descrizione, resa, varianti: [...] } }
    """
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    # Riga 7 = intestazione (1-indexed)
    headers = [cell.value for cell in ws[7]]
    col = {v: i for i, v in enumerate(headers) if v}

    products = {}
    for row in ws.iter_rows(min_row=8, values_only=True):
        cod  = row[col["Cod. Prodotto"]]
        nome = row[col["Prodotto"]]
        if not cod or not nome:
            continue
        # Pulizia del nome (può contenere \xa0 spazi non-breaking)
        nome = str(nome).replace("\xa0", " ").strip()

        if cod not in products:
            products[cod] = {
                "cod":           cod,
                "nome":          nome,
                "slug":          slugify(nome),
                "linea":         str(row[col["LINEA"]] or "").strip(),
                "categoria":     str(row[col["Categoria"]] or "").strip(),
                "sottocategoria":str(row[col["Sottocategoria"]] or "").strip(),
                "descrizione":   str(row[col["Descrizione prodotto"]] or "").strip(),
                "resa":          str(row[col["Resa"]] or "").strip(),
                "url_kerakoll":  f"{BASE_URL}/p/{slugify(nome)}",
                "pdf_scheda_tecnica": None,
                "varianti":      [],
            }

        variante = {
            "articolo":    str(row[col["Articolo"]] or "").strip(),
            "descrizione": str(row[col["DESCRIZIONE"]] or "").strip(),
            "tipo_conf":   str(row[col["tipo confezione"]] or "").strip(),
            "vol_conf":    row[col["Vol Conf."]],
            "um_conf":     str(row[col["conf. UM"]] or "").strip(),
            "prezzo":      row[col["Prezzo"]],
            "euro_um":     str(row[col["€/conf. UM"]] or "").strip(),
            "prezzo_ab":   row[col["Prezzo A+B"]],
            "note":        str(row[col["Note"]] or "").strip(),
        }
        products[cod]["varianti"].append(variante)

    return products


async def extract_pdf_link(page, product_url: str, product_name: str) -> str | None:
    """
    Naviga la pagina prodotto e tenta di estrarre il link PDF della scheda tecnica italiana.
    Strategia multi-step:
    1) Cerca link PDF diretti nella pagina
    2) Clicca "Scheda Tecnica" nel dropdown Downloads per espandere le lingue
    3) Cerca link con _ita o /ITALIA/ nel URL
    """
    try:
        resp = await page.goto(product_url, wait_until="domcontentloaded", timeout=PAGE_TIMEOUT)
        if resp is None or resp.status >= 400:
            print(f"    ✗ HTTP {resp.status if resp else '?'}: {product_url}")
            return None
    except PlaywrightTimeout:
        print(f"    ✗ Timeout: {product_url}")
        return None
    except Exception as e:
        print(f"    ✗ Errore navigazione: {e}")
        return None

    # Attendi che la pagina sia stabile
    await asyncio.sleep(1.5)

    # ── Strategia 1: Cerca link PDF italiani già visibili ─────────────────────
    pdf_url = await page.evaluate("""
        () => {
            const links = Array.from(document.querySelectorAll('a[href]'));
            
            const isTechSheet = (href) => {
                const lower = href.toLowerCase();
                return lower.includes('.pdf') && 
                       !lower.includes('security') && 
                       !lower.includes('hazex') &&
                       !lower.includes('techsiteguide') &&
                       !lower.includes('brochure');
            };

            // Priorità 1: link techSheet con ITALIA o _ita
            const best = links.find(a => 
                isTechSheet(a.href) && 
                a.href.includes('techSheet') && 
                (a.href.includes('_ita') || a.href.includes('ITALIA') || a.href.includes('/it/'))
            );
            if (best) return best.href;

            // Priorità 2: qualsiasi techSheet
            const tech = links.find(a => 
                isTechSheet(a.href) && 
                a.href.includes('techSheet')
            );
            if (tech) return tech.href;

            // Priorità 3: link con _ita o /ITALIA/ nel URL e .pdf
            const ita = links.find(a =>
                isTechSheet(a.href) &&
                (a.href.includes('_ita') || a.href.includes('ITALIA') || a.href.includes('/it/'))
            );
            if (ita) return ita.href;

            // Fallback: qualsiasi PDF della scheda tecnica
            const any = links.find(a =>
                isTechSheet(a.href) &&
                (a.href.includes('scheda') || a.href.includes('technical'))
            );
            if (any) return any.href;
            return null;
        }
    """)
    if pdf_url:
        return pdf_url

    # ── Strategia 2: Clicca sul trigger "Scheda Tecnica" nel dropdown ─────────
    try:
        # Trova il pulsante/accordion "Scheda Tecnica"
        trigger = await page.query_selector("text=Scheda Tecnica")
        if trigger:
            await trigger.click()
            await asyncio.sleep(1.5)

            # Ora cerca link con _ita o /ITALIA/ comparsi dopo il click
            pdf_url = await page.evaluate("""
                () => {
                    const links = Array.from(document.querySelectorAll('a[href]'));
                    
                    const isTechSheet = (href) => {
                        const lower = href.toLowerCase();
                        return lower.includes('.pdf') && 
                               !lower.includes('security') && 
                               !lower.includes('hazex') &&
                               !lower.includes('techsiteguide') &&
                               !lower.includes('brochure');
                    };

                    const best = links.find(a => 
                        isTechSheet(a.href) && 
                        a.href.includes('techSheet') && 
                        (a.href.includes('_ita') || a.href.includes('ITALIA') || a.href.includes('/it/'))
                    );
                    if (best) return best.href;

                    const tech = links.find(a => 
                        isTechSheet(a.href) && 
                        a.href.includes('techSheet')
                    );
                    if (tech) return tech.href;

                    const ita = links.find(a =>
                        isTechSheet(a.href) &&
                        (a.href.includes('_ita') || a.href.includes('ITALIA') || a.href.includes('/it/'))
                    );
                    if (ita) return ita.href;

                    const any = links.find(a =>
                        isTechSheet(a.href) &&
                        a.href.includes('media.kerakoll')
                    );
                    return any ? any.href : null;
                }
            """)
            if pdf_url:
                return pdf_url
    except Exception:
        pass

    # ── Strategia 3: Cerca intercettando le richieste di rete ─────────────────
    # (già fatto: non trovato niente)

    # ── Strategia 4: Cerca nel sorgente della pagina link PDF ─────────────────
    try:
        content = await page.content()
        # Cerca URL PDF nel sorgente HTML
        pdf_matches = re.findall(r'https://[^\s"\'<>]+\.pdf', content)
        ita_pdfs = [u for u in pdf_matches if '_ita' in u or 'ITALIA' in u.upper()]
        if ita_pdfs:
            return ita_pdfs[0]
        if pdf_matches:
            return pdf_matches[0]
    except Exception:
        pass

    return None


async def scrape_product(context, product: dict, semaphore: asyncio.Semaphore, idx: int, total: int) -> dict:
    """Scrapes un singolo prodotto usando una tab del browser."""
    async with semaphore:
        name = product["nome"]
        url  = product["url_kerakoll"]
        print(f"  [{idx+1}/{total}] {name} → {url}")

        page = await context.new_page()
        try:
            pdf = await extract_pdf_link(page, url, name)
            if pdf:
                product["pdf_scheda_tecnica"] = pdf
                print(f"    ✓ PDF: {pdf}")
            else:
                print(f"    ○ Nessun PDF trovato")
        finally:
            await page.close()

        return product


async def run_scraper():
    print("=" * 60)
    print("  Kerakoll Scraper — Schede Tecniche PDF")
    print("=" * 60)

    # ── Carica listino Excel ──────────────────────────────────────────────────
    print(f"\n[INFO] Caricamento listino Excel: {EXCEL_FILE}")
    products_dict = load_excel_products(EXCEL_FILE)
    products_list = list(products_dict.values())
    total = len(products_list)
    print(f"   {total} prodotti trovati nel listino\n")

    # ── Carica progress checkpoint se esiste ─────────────────────────────────
    already_done = {}
    progress_path = Path(PROGRESS_FILE)
    if progress_path.exists():
        with open(progress_path, "r", encoding="utf-8") as f:
            already_done = json.load(f)
        done_count = sum(1 for v in already_done.values() if v is not None)
        print(f"[CHECKPOINT] Trovato: {len(already_done)} prodotti già processati ({done_count} con PDF)\n")

    # Aggiorna i prodotti già processati
    for p in products_list:
        if p["cod"] in already_done:
            p["pdf_scheda_tecnica"] = already_done[p["cod"]]

    # Prodotti da scrapare (non ancora processati)
    to_scrape = [p for p in products_list if p["cod"] not in already_done]
    print(f"[SEARCH] Prodotti da scrapare: {len(to_scrape)}\n")

    if not to_scrape:
        print("[OK] Tutti i prodotti già processati!\n")
    else:
        # ── Avvia Playwright ──────────────────────────────────────────────────
        async with async_playwright() as pw:
            browser = await pw.chromium.launch(
                headless=True,
                args=["--no-sandbox", "--disable-dev-shm-usage"]
            )
            context = await browser.new_context(
                locale="it-IT",
                user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
                viewport={"width": 1280, "height": 800},
            )

            # Block risorse non necessarie (velocizza lo scraping)
            await context.route("**/*.{png,jpg,jpeg,gif,svg,woff,woff2,ttf,mp4,webp}", lambda r: r.abort())

            semaphore = asyncio.Semaphore(CONCURRENCY)
            tasks = [
                scrape_product(context, p, semaphore, i, len(to_scrape))
                for i, p in enumerate(to_scrape)
            ]

            # Esegui in batch, salvando il progress ogni 10 prodotti
            batch_size = 10
            for i in range(0, len(tasks), batch_size):
                batch = tasks[i:i + batch_size]
                results = await asyncio.gather(*batch)

                # Aggiorna checkpoint
                for p in results:
                    already_done[p["cod"]] = p["pdf_scheda_tecnica"]
                with open(progress_path, "w", encoding="utf-8") as f:
                    json.dump(already_done, f, ensure_ascii=False, indent=2)

                done = min(i + batch_size, len(to_scrape))
                pdf_found = sum(1 for v in already_done.values() if v)
                print(f"\n  [SAVE] Checkpoint salvato: {done}/{len(to_scrape)} — PDF trovati: {pdf_found}\n")

            await context.close()
            await browser.close()

    # ── Applica i risultati finali ai prodotti ────────────────────────────────
    for p in products_list:
        p["pdf_scheda_tecnica"] = already_done.get(p["cod"])

    # ── Salva output JSON ─────────────────────────────────────────────────────
    output_path = Path(OUTPUT_FILE)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Statistiche finali
    with_pdf    = sum(1 for p in products_list if p["pdf_scheda_tecnica"])
    without_pdf = total - with_pdf

    output = {
        "meta": {
            "source":    EXCEL_FILE,
            "listino":   "Novembre 2025",
            "total":     total,
            "with_pdf":  with_pdf,
            "without_pdf": without_pdf,
        },
        "products": products_list,
    }

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2, default=str)

    print("\n" + "=" * 60)
    print(f"  [COMPLETATO] Elaborazione terminata!")
    print(f"  Prodotti totali:      {total}")
    print(f"  Con scheda tecnica:   {with_pdf}")
    print(f"  Senza scheda tecnica: {without_pdf}")
    print(f"  Output JSON: {OUTPUT_FILE}")
    print("=" * 60)


if __name__ == "__main__":
    asyncio.run(run_scraper())
