# -*- coding: utf-8 -*-
"""Test veloce: scrapa solo 5 prodotti per verificare che il meccanismo funzioni."""

import asyncio
import json
import re
import sys
import unicodedata
import openpyxl
from playwright.async_api import async_playwright, TimeoutError as PlaywrightTimeout

sys.stdout.reconfigure(encoding="utf-8")

BASE_URL = "https://it.kerakoll.com"
PAGE_TIMEOUT = 25000

def slugify(text):
    if not text:
        return ""
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    text = text.lower()
    text = re.sub(r"['\u2019\u2018]", "", text)
    text = re.sub(r"[^a-z0-9\s\-]", " ", text)
    text = re.sub(r"\s+", "-", text.strip())
    text = re.sub(r"-+", "-", text)
    return text

async def extract_pdf(page, url):
    try:
        resp = await page.goto(url, wait_until="domcontentloaded", timeout=PAGE_TIMEOUT)
        if resp and resp.status >= 400:
            return None, f"HTTP {resp.status}"
    except PlaywrightTimeout:
        return None, "Timeout"
    except Exception as e:
        return None, str(e)

    await asyncio.sleep(2)

    # Strategia 1: link PDF diretti
    pdf_url = await page.evaluate("""
        () => {
            const links = Array.from(document.querySelectorAll('a[href]'));
            const ita = links.find(a =>
                a.href.toLowerCase().includes('.pdf') &&
                (a.href.includes('_ita') || a.href.toUpperCase().includes('ITALIA') || a.href.includes('/it/'))
            );
            if (ita) return ita.href;
            const any = links.find(a =>
                a.href.toLowerCase().includes('.pdf') &&
                (a.href.includes('techSheet') || a.href.includes('scheda') || a.href.includes('media.kerakoll'))
            );
            return any ? any.href : null;
        }
    """)
    if pdf_url:
        return pdf_url, "direct"

    # Strategia 2: clicca "Scheda Tecnica"
    try:
        trigger = page.locator("text=Scheda Tecnica").first
        if await trigger.count() > 0:
            await trigger.click()
            await asyncio.sleep(2)
            pdf_url = await page.evaluate("""
                () => {
                    const links = Array.from(document.querySelectorAll('a[href]'));
                    const ita = links.find(a =>
                        a.href.toLowerCase().includes('.pdf') &&
                        (a.href.includes('_ita') || a.href.toUpperCase().includes('ITALIA'))
                    );
                    if (ita) return ita.href;
                    const any = links.find(a =>
                        a.href.toLowerCase().includes('.pdf') &&
                        a.href.includes('media.kerakoll')
                    );
                    return any ? any.href : null;
                }
            """)
            if pdf_url:
                return pdf_url, "after-click"
    except Exception as e:
        pass

    # Strategia 3: cerca nel sorgente HTML
    try:
        content = await page.content()
        matches = re.findall(r'https://[^\s"\'<>]+\.pdf', content)
        ita = [u for u in matches if '_ita' in u or 'ITALIA' in u.upper()]
        if ita:
            return ita[0], "html-source-ita"
        if matches:
            return matches[0], "html-source"
    except Exception:
        pass

    return None, "not-found"

async def main():
    print("🧪 TEST SCRAPING — 5 prodotti campione\n")

    wb = openpyxl.load_workbook("kerakoll_listino-gen-ita-nov2025.xlsx")
    ws = wb.active
    headers = [cell.value for cell in ws[7]]
    col = {v: i for i, v in enumerate(headers) if v}

    # Prendi i primi 5 prodotti unici
    seen = set()
    test_products = []
    for row in ws.iter_rows(min_row=8, values_only=True):
        cod = row[col["Cod. Prodotto"]]
        nome = str(row[col["Prodotto"]] or "").replace("\xa0", " ").strip()
        if cod and nome and cod not in seen:
            seen.add(cod)
            test_products.append({"cod": cod, "nome": nome, "slug": slugify(nome)})
            if len(test_products) == 5:
                break

    async with async_playwright() as pw:
        browser = await pw.chromium.launch(headless=True, args=["--no-sandbox"])
        context = await browser.new_context(
            locale="it-IT",
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/124.0.0.0 Safari/537.36",
            viewport={"width": 1280, "height": 900},
        )
        await context.route("**/*.{png,jpg,jpeg,gif,svg,woff,woff2,ttf,mp4,webp,ico}", lambda r: r.abort())

        results = []
        for p in test_products:
            url = f"{BASE_URL}/p/{p['slug']}"
            print(f"  [{p['cod']}] {p['nome']}")
            print(f"  URL: {url}")
            page = await context.new_page()
            pdf, method = await extract_pdf(page, url)
            await page.close()
            status = f"✓ [{method}] {pdf}" if pdf else "✗ Nessun PDF"
            print(f"  → {status}\n")
            results.append({"cod": p["cod"], "nome": p["nome"], "url": url, "pdf": pdf, "method": method})

        await context.close()
        await browser.close()

    with open("test_scrape_results.json", "w", encoding="utf-8") as f:
        json.dump(results, f, indent=2, ensure_ascii=False)

    found = sum(1 for r in results if r["pdf"])
    print(f"📊 Risultato test: {found}/5 PDF trovati")
    print("📁 Dettagli salvati in test_scrape_results.json")

if __name__ == "__main__":
    asyncio.run(main())
